import os
import numpy as np
import pandas as pd
from typing import List
from datetime import datetime, timedelta
import pickle
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
import requests
from variables import *
from bs4 import BeautifulSoup
import re
from scipy.stats import norm
from utils import *
import re
import plotly.graph_objects as go
from pyhtml2pdf import  converter
from plotly.colors import n_colors
from bs4 import BeautifulSoup
from pytz import timezone    



def create_folder(path):
    # Create a foder in case it has not been created already
    if not os.path.exists(path):
        os.mkdir(path)
        print(f"Directory created: {path}")


#  Auxiliary functions for styling the table

def scale_col_range(col : pd.Series, range):
    """Scale column to range
    col = [0,10,100], range = 10 => [0,1,10]
    """
    return ((range) * ((col - col.min())/(col.max() - col.min() +1e-9))).astype(np.int32)

def bold(col:pd.Series)->List:
    """
    Return list string in col in bold using HTML syntax
    """
    return [f"<b>{element}</b>" for element in col]

def posneg_binary_color(col:pd.Series,pos_color,neg_color):
    """
    Return series with colors to differentiate positive and negative values
    """
    aux = col.reset_index()
    aux['color'] = neg_color #
    aux.loc[col>=0,'color'] = pos_color
    return np.array(aux['color'])

def posneg_gradient(col:pd.Series):
    aux = col.reset_index()
    aux['r'] = aux['g']  = aux['b'] = 255
    aux.loc[aux.Änderung<0,'g'] = aux.loc[aux.Änderung<0,'b'] = 255 - 150*aux.Änderung/(aux.Änderung.min()) 
    aux.loc[aux.Änderung>0,'g'] = aux.loc[aux.Änderung>0,'r'] = 255 - 150*aux.Änderung/(aux.Änderung.max())
    aux = aux.astype(np.int32)
    return  np.array([f"rgb({aux.loc[i,'r']},{aux.loc[i,'g']},{aux.loc[i,'b']})" for i in range(aux.shape[0]) ])

#################################################################################################

def save_as_pickle(variable, path):
    with open(path, 'wb') as handle:
        pickle.dump(variable, handle, protocol=pickle.HIGHEST_PROTOCOL)

def read_pickle(path):
    with open(path, 'rb') as handle:
        variable = pickle.load(handle)
        return variable
    


def next_business_day(input_datetime):
    # Define a list of weekdays that are not considered business days (Saturday and Sunday)
    non_business_days = [5, 6]  # 5 represents Saturday, and 6 represents Sunday

    # Start with the next day from the input date
    next_day = input_datetime + timedelta(days=1)

    # Check if the next day is a non-business day (Saturday or Sunday)
    while next_day.weekday() in non_business_days:
        next_day += timedelta(days=1)

    return next_day





def get_overview(auswahl, tries=10):
    print(f"################## Obtaining overview ##################")
    ## Download tradingDates and future_date_col
    if auswahl == 1:
    # STOXX
        url = 'https://www.eurex.com/api/v1/overallstatistics/69660'
        print("STOXX")
    else:
    # DAX
        url = 'https://www.eurex.com/api/v1/overallstatistics/70044'
        print ("DAX")
        
    headers = {'Accept': '*/*'}
    params_ov = {'filtertype': 'overview', 'contracttype': 'M'}
    print('Getting data from EUREX...')

    for i in range(tries):
        print(f'try = {i}')
        try:
            response = requests.get(url, params=params_ov, headers=headers,timeout = 8)

            if not response.ok:
                raise ValueError("Conection failed.")

            tradingDates = pd.to_datetime(pd.Series(response.json()['header']['tradingDates']),format="%d-%m-%Y %H:%M")

            #print("Index     Date")
            #for idx,date in enumerate(tradingDates.dt.strftime("%d-%m-%Y")):
            #    print(f" {idx:<6} {date:^3}")

            #date_idxs = get_date_idx(tradingDates.index)
            params_ov['busdate'] = tradingDates[0].strftime("%Y%m%d")
            response = requests.get(url, params=params_ov, headers=headers,timeout = 8)
            if not response.ok:
                    raise ValueError("Conection failed.")

            params_details = {}
            if 'dataRows' in response.json(): 
                overview_df = pd.DataFrame(response.json()['dataRows'])
                overview_df = overview_df[overview_df.contractType == 'M']
                overview_df.sort_values('date',ignore_index=True,inplace=True)
                future_date_col = overview_df.loc[:,'date']
                overview_df.date = pd.to_datetime(overview_df.date,format='%Y%m%d').dt.strftime('%d-%m-%Y')

            else:
                raise ValueError

            params_details['busdate'] = f"{tradingDates[1].strftime('%Y%m%d')}"
            response = requests.get(url, params=params_details, headers=headers,timeout = 8)
            if not response.ok:
                raise ValueError("Conection failed.")
            break
        except:
            continue
            
    return url, headers, tradingDates, future_date_col, overview_df


def get_contracts(heute, url, headers, tradingDates, future_date_col,tries=10):
    #https://en.wikipedia.org/wiki/Offset_(computer_science)#
    offset = 0
    # **tbd: heute shall be a working day (Mo - Fr)**
    expiry = datetime.strptime(future_date_col[0],"%Y%m%d")
    expiry_1 = datetime.strptime(future_date_col[1],"%Y%m%d")

    tage_bis_verfall = (expiry - heute).days
    
    if tage_bis_verfall < 0:
        offset = 1

    expiry = datetime.strptime(future_date_col[0 + offset],"%Y%m%d")
    expiry_1 = datetime.strptime(future_date_col[1 + offset],"%Y%m%d") 
    tage_bis_verfall = (expiry - heute).days +1

    print(f"################## Obtaining contracts ##################")
    params_details = {}
    params_details['filtertype'] = 'detail'
    params_details['contracttype'] = 'M'
    dict_prod_bus = {}
    for productdate_idx in [0,1]:
        dict_prod_bus[productdate_idx] = {}
        params_details['productdate'] = future_date_col[productdate_idx + offset]
        for busdate_idx in [0,1]:
            params_details['busdate'] = f"{tradingDates[busdate_idx].strftime('%Y%m%d')}" 
            for i in range(tries):
                print(f"busdate_idx = {busdate_idx}   | productdate_idx = {productdate_idx} | try = {i}")
                try:
                    if (busdate_idx == 1) and (productdate_idx == 1):
                        # No need to request data for this case so we skip this iteration
                        break
                    
                    # Request data
                    response = requests.get(url, params=params_details, headers=headers, timeout = 10)
                    if not response.ok:
                        raise ValueError("Conection failed.")
                    contractsCall_aux_df =  pd.DataFrame(response.json()['dataRowsCall'])
                    contractsPut_aux_df =  pd.DataFrame(response.json()['dataRowsPut'])
                    
                    # Create dataframe with the just requested data

                    aux_df = contractsCall_aux_df[['strike','openInterest']].merge(
                        contractsPut_aux_df[['strike','openInterest']],
                        on = "strike",
                        how="left",
                        suffixes=('_CF', '_PF')
                    )
                    dict_prod_bus[productdate_idx][busdate_idx] = aux_df
                except:
                    continue
                break
    return dict_prod_bus, expiry, expiry_1, tage_bis_verfall


def get_date_idx(list_idxs)->list:
    while(True):
        date_idxs = [0, 1]
        date_idxs = list(map(int,date_idxs))
        date_idxs = [idx for idx in date_idxs if idx in list_idxs]
        if len(date_idxs) < 1:
            print("No valid indexes passed.")
        else:
            return date_idxs
        

def get_heute():
    heute = datetime.now(timezone('Europe/Berlin')).replace(tzinfo=None)
    if heute.weekday() > 4:
        #sunday => wd = 6 -> td = +1
        #sartuday => wd = 5 > td +2
        heute += timedelta(days= 7 - heute.weekday())
    return heute



def get_euribor_3m()->float:
    euribor_3m_df = pd.read_html("https://www.euribor-rates.eu/en/current-euribor-rates/2/euribor-rate-3-months/")[0].rename(columns= {0: "Date", 1 : "InterestRate"})
    euribor_row = euribor_3m_df.loc[0]
    print(f"Euribor 3M in {euribor_row['Date']} is {euribor_row['InterestRate']}")
    return float(euribor_row["InterestRate"].replace("%",""))/100

def get_finazen_price(stock_idx)->float:
    dict_stock_names = {
    0 : 'vdax-new-2m',
    1 : 'vstoxx'
    }
    stock = dict_stock_names[stock_idx]
    #accepted values for stock: , 
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36'
    }
    response = requests.get('https://www.finanzen.net/index/' + stock, headers=headers)
    html = BeautifulSoup(response.content, 'html.parser')
    aux = BeautifulSoup.findAll(html,id = 'snapshot-value-fst-current-0')[0].span.get_text()
    price = re.findall(r'\d{0,},\d{0,}',aux)[0]
    price = float(price.replace(',','.'))
    print(f"{stock} = {price} PTS")
    return price



def is_calculation_needed(auswahl, tage_bis_verfall):
    dict_condition = {
    "always" : True,
    "tage_bis_verfall<=5" : tage_bis_verfall < 5
    }
    list_email_send = []
    for email in list_emails:
        if (dict_index_stock[auswahl] == email['index']) and (dict_condition[email['condition']]):
            list_email_send+= [email['id']]
    return list_email_send

def parse_eurex(auswahl, heute = None):
    """
    auswahl: Option defined by the user. See variables.dict_index_stock.
    heute: it should be a date as string using dd/mm/yyyy format. If None, it will be defined by get_heute().
    """

    if not heute:
        heute = get_heute()
    else:
        heute = datetime.strptime(heute,"%d/%m/%Y")

    print(f"heute = {heute}")

    # First stage
    url, headers, tradingDates, future_date_col, overview_df = get_overview(auswahl)
    dict_prod_bus, expiry, expiry_1, tage_bis_verfall =  get_contracts(heute, url, headers, tradingDates, future_date_col)
    
    list_email_send_selection  = is_calculation_needed(auswahl, tage_bis_verfall)

    if len(list_email_send_selection) == 0:
        print(f'Files for {dict_index_stock[auswahl]} were not genereted.')
        return None

    stocks = pd.read_html("https://www.boerse-stuttgart.de/en/")

    if auswahl == 0: # DAX
        stocks_df = stocks[0]
        stock_price = stocks_df.loc[stocks_df['Indices GER'] == "L&S DAX","Price"].values[0]


    else: # STOXX
        stocks_df = stocks[1]
        stock_price = stocks_df.loc[stocks_df['Indices EU / USA / INT'] == "CITI Euro Stoxx 50","Price"].values[0]

    Spannweite, *_ = get_default_values(auswahl).values()
    span = (Spannweite/4)
    ZentralKurs = round(stock_price/span)*span

    nbd_heute = next_business_day(tradingDates[0])
    nbd_last = tradingDates[0]
    nbd_dict = {'heute' : nbd_heute, 'last' : nbd_last}

    InterestRate = get_euribor_3m()
    volatility = get_finazen_price(auswahl)/100

    print(f'heute = {heute.strftime("%d/%m/%Y")}')
    print(f"tage_bis_verfall = {tage_bis_verfall}")
    print(f"expiry = {expiry}")
    print(f"expiry_1 = {expiry_1}")
    print(f"stock_price = {stock_price}")
    print(f"ZentralKurs = {ZentralKurs}")
    print(f"volatility = {volatility}")
    print(f"InterestRate = {InterestRate}")

    return auswahl, ZentralKurs, volatility, InterestRate, tage_bis_verfall, nbd_dict, dict_prod_bus, stock_price, expiry, expiry_1, heute, list_email_send_selection, future_date_col

def hedge(auswahl, ZentralKurs, volatility, InterestRate, tage_bis_verfall, dict_prod_bus, stock_price, expiry, expiry_1, heute, export_excel = False):

    Spannweite, Schritt,volatility_Laufzeit,KontraktWert = get_default_values(auswahl).values()

    Minkurs = ZentralKurs - (Spannweite/ 2)
    Maxkurs = Minkurs + Spannweite
    Schritte = int(Spannweite / Schritt)

    DetailMin = round(stock_price - (Spannweite)/4)
    DetailMax = DetailMin + (Spannweite/2)


    if tage_bis_verfall >= 1:
        delta = 0.5
    else:
        delta = 1

    print(f"Minkurs = {Minkurs}")
    print(f"Maxkurs = {Maxkurs}")
    print(f"Schritte = {Schritte}")
    print(f"volatility_Laufzeit = {volatility_Laufzeit}")


    # Second stage
    # Create Basis column that will be used for multiple tables
    Basis = pd.Series((Minkurs + np.arange(int(Schritte) +1)* Schritt)[::-1])

    # Transform series into dataframe to make dataframe creation easier 
    Basis_df = Basis.reset_index().rename(columns={0:"Basis"})[["Basis"]]


    # Ueberhaenge_df, Summery_df are created from a copy of the same Dataframe
    # so they have the same index. It means they could be in one  Dataframe but
    #  I decided to keep them separated as they were in the VBA code.

    Ueberhaenge_df = Basis_df.copy()
    Summery_df = Basis_df.copy()

    for productdate_idx in [0,1]:
        for busdate_idx in [0,1]:

            if (busdate_idx == 1) and (productdate_idx == 1):
                continue

            # Create dataframe with the just requested data
            aux_df = Basis_df.copy()

            aux_df = aux_df.merge(
                dict_prod_bus[productdate_idx][busdate_idx],
                left_on = "Basis",
                right_on = "strike",
                how= "left"
            )

            if productdate_idx == 0:
                Ueberhaenge_df[busdate_idx] = aux_df.openInterest_PF - aux_df.openInterest_CF

                if busdate_idx == 0:
                    Summery_df["openInterest_PF"] = aux_df["openInterest_PF"]
                    Summery_df["openInterest_CF"] = aux_df["openInterest_CF"]
            
            elif productdate_idx == 1:
                if busdate_idx == 0:
                    Ueberhaenge_df["nextContract"] = -(aux_df.openInterest_PF - aux_df.openInterest_CF)
                    Summery_df["nextContract"] = Ueberhaenge_df["nextContract"]


    Ueberhaenge_df.rename(columns = {0 : "Front"},inplace=True)
    Ueberhaenge_df.rename(columns = {1 : "Last"},inplace=True)


    Summery_df["heute"] = Ueberhaenge_df["Front"] * (1 / KontraktWert) * delta
    Summery_df["last_day"] = Ueberhaenge_df["Last"] * (1 / KontraktWert) * delta

    Ueberhaenge_df["Summe"] = Ueberhaenge_df[["Front", "nextContract"]].sum(axis=1)
    Ueberhaenge_df = Ueberhaenge_df[['Basis', 'Summe',"Last",'Front', "nextContract"]]


    Summery_df["Änderung"] = Summery_df.heute - Summery_df.last_day

    if (delta == 1):
        Summery_df['nextContract'] =  Summery_df['nextContract'] / 2
    
    SummeryDetail_df = Summery_df[(Summery_df.Basis >= DetailMin) & (Summery_df.Basis < (DetailMax + Schritt))]
    
    # Third stage
    SchrittWeite = 10



    Tage = tage_bis_verfall
    if Tage == 0:
        Tage = 0.5

    Tage_1 = (expiry_1 - heute).days # ASK: Add 1?

    print(f"Tage = {Tage}")
    print(f"Tage_1 = {Tage_1}")

    Kurs_count = int(Spannweite/SchrittWeite)+1 
    HedgeBedarf_kurs=  pd.DataFrame(Maxkurs - np.arange(Kurs_count)* SchrittWeite,columns= ["Basis"])


    Hedge_dimensions = Kurs_count,int(Schritte+1)
    HedgeBedarf_values = np.zeros(Hedge_dimensions)
    HedgeBedarf1_values = np.zeros(Hedge_dimensions)


    for k in range(Hedge_dimensions[1]):
        Basis_value = Basis[k]
        Kontrakte = Ueberhaenge_df.loc[k,"Front"]
        Kontrakte_1 = Ueberhaenge_df.loc[k,"nextContract"] # ASK: Why negative?
        for i in range(Hedge_dimensions[0]):
            Kurs = Maxkurs - i * SchrittWeite
            # In python np.log = natural log
            h1 = np.log(Kurs / Basis_value)
            if auswahl == 0:
                sigma = volatility * ((Tage / volatility_Laufzeit) ** 0.5)
            else:
                sigma = volatility

            sigma_1 = volatility * ((Tage_1 / volatility_Laufzeit) ** 0.5)
            h2 = InterestRate + sigma * sigma / 2
            h2_1 = InterestRate + sigma_1 * sigma_1 / 2
            d1 = (h1 + (h2 * (Tage / 365))) / (sigma * ((Tage / 365) ** 0.5))
            d1_1 = (h1 + (h2_1 * (Tage_1 / 365))) / (sigma_1 * ((Tage_1 / 365) ** 0.5))
            Phi = norm.pdf(d1, 0, 1)               
            Phi_1 = norm.pdf(d1_1, 0, 1)
            Gamma = Phi / (Kurs * (sigma * (Tage / 365) ** 0.5))
            Gamma_1 = Phi_1 / (Kurs * (sigma_1 * (Tage_1 / 365) ** 0.5))
            HedgeBedarf_values[i,k] = Gamma * Kontrakte / KontraktWert
            HedgeBedarf1_values[i,k] = Gamma_1 * Kontrakte_1 / KontraktWert

            if Kurs == 17000:
                continue
    HedgeSum = HedgeBedarf_values.sum(axis=1)/2
    HedgeSum_1 = HedgeBedarf1_values.sum(axis=1)/2

    HedgeBedarf_df = pd.DataFrame(data =HedgeBedarf_values, columns= Basis )
    HedgeSum_df = pd.DataFrame(HedgeSum,columns=["HedgeSum"])
    HedgeBedarf_df = pd.concat([HedgeBedarf_kurs,HedgeSum_df,HedgeBedarf_df], axis=1)

    HedgeBedarf1_df = pd.DataFrame(data =HedgeBedarf1_values, columns= Basis)
    HedgeSum1_df = pd.DataFrame(HedgeSum_1,columns=["HedgeSum"])
    HedgeBedarf1_df = pd.concat([HedgeBedarf_kurs,HedgeSum1_df,HedgeBedarf1_df], axis=1)


    # Fourth Stage
    info_df = pd.DataFrame({
        'ZentralKurs' : [ZentralKurs],
        'Spannweite': [Spannweite],
        'SchrittWeite': [SchrittWeite],
        'Schritt' : [Schritt],
        'auswahl': dict_index_stock[auswahl],
        'expiry': expiry,
        'expiry_1': expiry_1,
        'heute': heute,
        'tage_bis_verfall': tage_bis_verfall,
        'stock_price': stock_price,
        'Minkurs': Minkurs,
        'Maxkurs': Maxkurs,
        'Schritte': Schritte,
        'DetailMin': DetailMin,
        'DetailMax': DetailMax,
        'delta' : delta,
        'volatility' : volatility
    }).T.reset_index().rename(columns={0:"Value", 'index': "Info"})

    if export_excel:
        list_excel_files = [
        os.path.join(current_results_path, f"{dict_index_stock[auswahl]}.xlsx"),
        os.path.join(old_results_path, f"{dict_index_stock[auswahl]}_{heute.strftime('%Y_%m_%d')}.xlsx")
        ]

        for excel_file in list_excel_files:
            with pd.ExcelWriter(excel_file,datetime_format="DD/MM/YYYY") as writer:
                info_df.to_excel(writer,sheet_name='infos',index=False)
                Ueberhaenge_df.to_excel(writer,sheet_name='Ueberhaenge',index=False)
                Summery_df.to_excel(writer,sheet_name='Summery',index=False)
                SummeryDetail_df.to_excel(writer,sheet_name='SummeryDetail',index=False)
                HedgeBedarf_df.to_excel(writer,sheet_name='HedgeBedarf',index=False)
                HedgeBedarf1_df.to_excel(writer,sheet_name='HedgeBedarf+01',index=False)
                
        print("Excel files have been exported.")

    return Summery_df, HedgeBedarf_df, HedgeBedarf1_df, Ueberhaenge_df, delta

def parse_excel(auswahl : int, excel_path : str):

    dict_auswahl_prefix = {
        0 : "",
        1 : "STOXX_"
    }

    wb = load_workbook(excel_path, data_only=True, keep_vba=True)
    Ueberhaenge_sheet = wb['Ueberhaenge']
    heute = Ueberhaenge_sheet.cell(*coordinate_to_tuple("G5")).value
    expiry = Ueberhaenge_sheet.cell(*coordinate_to_tuple("G3")).value
    expiry_1 = Ueberhaenge_sheet.cell(*coordinate_to_tuple("R6")).value
    stock_price = Ueberhaenge_sheet.cell(*coordinate_to_tuple("X4")).value
    InterestRate = Ueberhaenge_sheet.cell(*coordinate_to_tuple("N3")).value
    ZentralKurs =  Ueberhaenge_sheet.cell(*coordinate_to_tuple("C3")).value  
    volatility = Ueberhaenge_sheet.cell(*coordinate_to_tuple("N4")).value

    future_date_col = [expiry, expiry_1]

    Summery_sheet = wb[f'{dict_auswahl_prefix[auswahl]}Summery']
    nbd_heute = Summery_sheet.cell(*coordinate_to_tuple("C9")).value
    
    nbd_last = Summery_sheet.cell(*coordinate_to_tuple("C9")).value
    nbd_last = next_business_day(heute)
    nbd_dict = {'heute' : nbd_heute, 'last' : nbd_last}


    tage_bis_verfall = (expiry - heute).days #+ 1

    list_email_send_selection  =  is_calculation_needed(auswahl, tage_bis_verfall)

    dict_prod_bus = {}
    for productdate_idx in [0,1]:
        dict_prod_bus[productdate_idx] = {}
        for busdate_idx in [0,1]:
            print(f"busdate_idx = {busdate_idx}   | productdate_idx = {productdate_idx}")
            if (busdate_idx == 1) and (productdate_idx == 1):
                # No need to request data for this case so we skip this iteration
                break

            if (busdate_idx == 0) and (productdate_idx == 0):
                contractsCall_aux_df =  pd.read_excel(excel_path, sheet_name = f"{dict_auswahl_prefix[auswahl]}Call_Front")
                contractsPut_aux_df =  pd.read_excel(excel_path, sheet_name = f"{dict_auswahl_prefix[auswahl]}Put_Front")

            if (busdate_idx == 1) and (productdate_idx == 0):
                contractsCall_aux_df= pd.read_excel(excel_path, sheet_name = f"{dict_auswahl_prefix[auswahl]}CallFront-1")
                contractsPut_aux_df = pd.read_excel(excel_path, sheet_name = f"{dict_auswahl_prefix[auswahl]}PutFront-1")

            if (busdate_idx == 0) and (productdate_idx == 1):
                contractsCall_aux_df = pd.read_excel(excel_path, sheet_name = f"{dict_auswahl_prefix[auswahl]}Put+01")
                contractsPut_aux_df= pd.read_excel(excel_path, sheet_name = f"{dict_auswahl_prefix[auswahl]}Call+01")


            aux_df = contractsCall_aux_df[['strike','openInterest']].merge(
                contractsPut_aux_df[['strike','openInterest']],
                on = "strike",
                how="left",
                suffixes=('_CF', '_PF')
            )
            dict_prod_bus[productdate_idx][busdate_idx] = aux_df

    print(f'heute = {heute.strftime("%d/%m/%Y")}')
    print(f"tage_bis_verfall = {tage_bis_verfall}")
    print(f"expiry = {expiry}")
    print(f"expiry_1 = {expiry_1}")
    print(f"stock_price = {stock_price}")
    print(f"ZentralKurs = {ZentralKurs}")
    print(f"volatility = {volatility}")
    print(f"InterestRate = {InterestRate}")

    return auswahl, ZentralKurs, volatility, InterestRate, tage_bis_verfall, nbd_dict, dict_prod_bus, stock_price, expiry, expiry_1, heute, list_email_send_selection, future_date_col

def generate_pdfs(auswahl, Summery_df, HedgeBedarf_df, HedgeBedarf1_df, stock_price, heute, nbd_dict, tage_bis_verfall, delta, expiry, expiry_1, file_path):

    # Fifth stage
    # Values to create arrow
    idx_closest = (HedgeBedarf_df.Basis - stock_price).abs().idxmin()
    closest_Basis =  HedgeBedarf_df.loc[idx_closest,"Basis"]
    hedge_sum = HedgeBedarf_df.HedgeSum + HedgeBedarf1_df.HedgeSum
    closest_Sum = hedge_sum.loc[idx_closest]

    x_axis_length = hedge_sum.max() - hedge_sum.min()
    y_axis_length = HedgeBedarf_df.Basis.max() - HedgeBedarf_df.Basis.min()

    min_Kontrakte = 5000
    max_Differenz = 1000
    prozentual = 0.2

    dict_verfall_sufix= {
        True : "_verfall",
        False : ""
    }

    is_close_verfall =  tage_bis_verfall < 5
    for is_detailed in [False, True]:
        if is_detailed:
            indexes = Summery_df.loc[(Summery_df.Basis >= stock_price)].tail(10).index.to_list() + Summery_df.loc[(Summery_df.Basis < stock_price)].head(10).index.to_list()
            Summery_df =  Summery_df.loc[indexes].reset_index(drop=True)

            indexes = HedgeBedarf_df.loc[(HedgeBedarf_df.Basis >= Summery_df.Basis.min()) & (HedgeBedarf_df.Basis <= Summery_df.Basis.max())].index
            
            HedgeBedarf_df = HedgeBedarf_df.loc[indexes].reset_index(drop=True)
            HedgeBedarf1_df = HedgeBedarf1_df.loc[indexes].reset_index(drop=True)

        # (CF > min_Kontrakte) AND (PF > min_Kontrakte) AND (ABS(CF - PF)  < MIN(PF,CF)*)
        mask_highlights = (
            (Summery_df.openInterest_CF > min_Kontrakte) & 
            (Summery_df.openInterest_PF > min_Kontrakte) & 
            (
                (Summery_df.openInterest_CF - Summery_df.openInterest_PF).abs() < 
                (Summery_df[['openInterest_PF',"openInterest_CF"]].min(axis=1)*prozentual)))


        ################################## Hilights for Basis column ####################################

        Summery_df["basis_color"] = "lavender"
        if not is_detailed:
            Summery_df.loc[
                mask_highlights, "basis_color"] = "yellow"
        col_basis_color = Summery_df.basis_color.to_numpy()
        #################################################################################################

        ######################### Gradient of red and blue for Anderung column ##########################
        aux = Summery_df.Änderung.reset_index()
        aux['r'] = aux['g']  = aux['b'] = 255
        aux.loc[aux.Änderung<0,'g'] = aux.loc[aux.Änderung<0,'b'] = 255 - 150*aux.Änderung/(aux.Änderung.min()) 
        aux.loc[aux.Änderung>0,'g'] = aux.loc[aux.Änderung>0,'r'] = 255 - 150*aux.Änderung/(aux.Änderung.max())
        aux = aux.astype(float)
        rb_shades = np.array([f"rgb({aux.loc[i,'r']},{aux.loc[i,'g']},{aux.loc[i,'b']})" for i in range(aux.shape[0]) ])
        col_anderung_color = rb_shades
        #################################################################################################

        ############################# Colors for heute and last_day columns #############################
        col_heute_color = posneg_binary_color(Summery_df.heute,"rgb(0, 204, 204)","rgb(77, 166, 255)")
        col_last_day_color = posneg_binary_color(Summery_df.last_day,"rgb(0, 204, 204)","rgb(77, 166, 255)")
        #################################################################################################

        ###################### Gradient of green and blue for Put anc Call columns ######################
        col_put_color  = np.array(n_colors('rgb(214, 245, 214)', 'rgb(40, 164, 40)',
            20, colortype='rgb'))[scale_col_range(Summery_df.openInterest_PF,19)]

        col_call_color  = np.array(n_colors('rgb(204, 224, 255)', 'rgb(0, 90, 179)',
            20, colortype='rgb'))[scale_col_range(Summery_df.openInterest_CF,19)]
        #################################################################################################

        num_rows = Summery_df.shape[0]+1
        height = 1050 #row_height*42
        row_height = int(height/(num_rows))

        # width of the image
        widht = 1000
        row_height_percent =  (row_height/height)


        values_body = [
                    (Summery_df.Basis +1e-6).round(0).astype(int),
                    (Summery_df.Änderung +1e-6).round(0).astype(int),
                    (Summery_df.heute +1e-6).round(0).astype(int),
                    (Summery_df.last_day +1e-6).round(0).astype(int),
        ]

        values_header = bold(["Basis","Änderung",nbd_dict['heute'].strftime("%d/%m/%y"),nbd_dict['last'].strftime("%d/%m/%y")])


        if not  is_detailed:
            values_body += [(Summery_df.openInterest_PF), (Summery_df.openInterest_CF)]
            values_header += bold(["Put","Call"])


        font_size =  int(440/num_rows)

        fig = go.Figure(
            data=[
                go.Table(

                    # Define some paremeters for the header
                    header=dict(
                        # Names of the columns
                        values= values_header,
                        
                        # Header style
                        fill_color='paleturquoise',
                            align='center',
                            font = {'size': int(font_size*0.8)},
                            height = row_height,
                    ),

                    cells=dict(

                        align='center',
                        height = row_height,
                        font = {'size': font_size,},

                        # Values of the table
                        values= values_body,

                        # Colors of the columns
                        fill_color = [
                            col_basis_color,
                            col_anderung_color,
                            col_heute_color,
                            col_last_day_color,
                            col_put_color, 
                            col_call_color
                        ],
                    )
                )
            ],
        )
        fig.update_layout(
            height=1050, width=550,
            margin=dict(l=0,r=0,b=0.0,t=0)
        )

        fig.write_image(os.path.join(current_results_path,f'image_table{dict_verfall_sufix[is_detailed]}.svg'),scale=1)


        data =  [

            # Create 0  y axis
            go.Scatter(
                x = [0,0], 
                y = [HedgeBedarf_df.Basis.min(), HedgeBedarf_df.Basis.max()],
                mode = "lines",
                marker_color = "orange",
                showlegend = False
            ),

            go.Scatter(
                x = hedge_sum,
                y = HedgeBedarf_df.Basis,
                mode = "lines",
                name =expiry.strftime("%Y-%m") + " + "+ expiry_1.strftime("%Y-%m"),
                marker_color= "blue"
            ),

        go.Scatter(
            x = [closest_Sum + x_axis_length/5,closest_Sum + x_axis_length/50], 
            y = [closest_Basis,closest_Basis],
            marker= dict(size=20,symbol= "arrow-bar-up", angleref="previous"),
            marker_color = "red",
            showlegend = False
        )
        ]
        
        min_x_list = [hedge_sum.min()]
        max_x_list = [hedge_sum.max()]
        if not is_detailed:
            min_x_list.append(HedgeBedarf1_df.HedgeSum.min())
            max_x_list.append(HedgeBedarf1_df.HedgeSum.max())
            data += [
                go.Scatter(
                    x = HedgeBedarf1_df.HedgeSum,
                    y = HedgeBedarf_df.Basis,
                    mode = "lines",
                    name = expiry_1.strftime("%Y-%m"),
                    marker_color= "rgb(255,0,255)"
                ),
            ]

        ##################################################################################################
        dx =  0.1*(hedge_sum.max() - hedge_sum.min())



        fig = go.Figure(data=data)

        #mask = HedgeBedarf_df.index % round(HedgeBedarf_df.shape[0]/40) == 0
        fig.update_layout(
            margin=dict(l=0,r=0,b=0.1,t=row_height),
            # Set limits in the x and y axis
            yaxis_range= [HedgeBedarf_df.Basis.min(), HedgeBedarf_df.Basis.max()],
            xaxis_range= [min(min_x_list) - dx, max(max_x_list) + dx],
            yaxis = dict(
                tickmode = 'array',
                tickvals = Summery_df.Basis, #HedgeBedarf_df.Basis[mask],
                ticktext = Summery_df.Basis,#HedgeBedarf_df.Basis[mask]
            ),
            
            # Remove margins
            paper_bgcolor="white",
            
            # Define width and height of the image
            width=380,height=1050,
            template = "seaborn",

            # Legend parameters
            legend=dict(
                yanchor="top",
                y=1,# - (row_height/height),
                xanchor="right",
                x= 1,
                font=dict(
                    size = 8
                ),
                # legend in the vertical
                orientation = "v",
                bgcolor  = 'rgba(0,0,0,0)'
                ),
        )


        fig.write_image(os.path.join(current_results_path,f'image_graph{dict_verfall_sufix[is_detailed]}.svg'),scale=1)



        fig = go.Figure()

        if not is_detailed:
            fig.add_trace(
                trace = go.Bar(name='Put', y=Summery_df.Basis, x=-Summery_df.openInterest_PF,orientation='h', marker_color = 'rgb(40, 164, 40)'),
            )

            fig.add_trace(
                trace = go.Bar(name='Call', y=Summery_df.Basis, x=Summery_df.openInterest_CF,orientation='h', marker_color = 'rgb(0, 90, 179)'),
            )



        fig.update_layout(
            #barmode='stack', 
            barmode='overlay', 
            margin=dict(l=0,r=0,b=0.1,t=row_height),
            # Set limits in the x and y axis
            #yaxis_range= [HedgeBedarf_df.Basis.min(), HedgeBedarf_df.Basis.max()],
            #xaxis_range= [hedge_sum.min() - dx, hedge_sum.max() + dx],
            
            # Define width and height of the image
            width=80,height=1050,
            template = "seaborn",
            showlegend=False,
            bargap =0.5 ,
            plot_bgcolor='rgba(0, 0, 0, 0)',
            paper_bgcolor='rgba(0, 0, 0, 0)',
        )


        fig.update_xaxes(visible=False)
        fig.update_yaxes(visible=False)

        fig.write_image(os.path.join(current_results_path,f'image_bar{dict_verfall_sufix[is_detailed]}.svg'),scale=1)


        with open(src_html) as file:
            template = file.read()

        if not is_close_verfall:
            subst = ""
            regex = r"\$BEGIN_DETAIL\$.*\$END_DETAIL\$"
            template = re.sub(regex, subst, template, 0, re.MULTILINE | re.DOTALL)
        else:
            template = template.replace("$BEGIN_DETAIL$", "").replace("$END_DETAIL$", "")

        dict_title = {
            0 : "OpenInterest und HedgeBedarf",
            1 : "STOXX 50 OpenInterest und HedgeBedarf",
        }

        # Replace specific characters in the template by values
        dict_raplace = {
            "$PUT_SUM$": int(Summery_df.openInterest_PF.sum()),
            "$CALL_SUM$": int(Summery_df.openInterest_CF.sum()),
            "$TBF$": tage_bis_verfall,
            "$DELTA$":str(delta).replace(".",","),
            "$DATE$": heute.strftime("%d/%m/%Y"),
            "$FRONT_DATE$" : expiry.strftime("%Y-%m"),
            "$TITLE$" : dict_title[auswahl],
            "$HEADER_COLOR$" : "background-color: rgb(12, 89, 177)"
        }

        for key, value in dict_raplace.items():
            template = template.replace(key, str(value))


        result_html = {True:summery_html, False: summery_verfall_html}[is_detailed]

        # Export html file
        with open(result_html,'w') as file:
            file.write(template)

        # Results files
        
    dict_auswahl_colors = {
        0 : {"$HEADER_COLOR$":"", "$FUTURE_COLOR$":"background-color: rgb(0, 174, 255);","$FONT_COLOR$":"color: black;"},
        1 : {"$HEADER_COLOR$":"background-color: rgb(12, 89, 177);", "$FUTURE_COLOR$":"", "$FONT_COLOR$" : "color: white;"},
    }

    with open(src_css) as file:
        css_file = file.read()

    for key, value in dict_auswahl_colors[auswahl].items():
        css_file = css_file.replace(key, str(value))

    with open(result_css,'w') as file:
        file.write(css_file)

    converter.convert("file://" + os.path.join(os.getcwd(),result_html), file_path)
    #shutil.copyfile(list_pdf_files[0], list_pdf_files[1])
    
    print("PDF has been generated.")